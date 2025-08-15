from fpdf import FPDF
import openpyxl
from datetime import datetime
from database import conectar_db, cerrar_db

def generar_reporte_usuario(usuario_id):
    conexion = conectar_db()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("SELECT * FROM usuarios WHERE id = %s", (usuario_id,))
    usuario = cursor.fetchone()
    if not usuario:
        cerrar_db(conexion)
        return None, None

    cursor.execute("""
        SELECT a.nombre, r.cantidad, r.calorias_totales, r.fecha, a.es_por_peso
        FROM registros_diarios r
        JOIN alimentos a ON r.alimento_id = a.id
        WHERE r.usuario_id = %s
        ORDER BY r.fecha DESC
        LIMIT 15
    """, (usuario_id,))
    alimentos = cursor.fetchall()
    cerrar_db(conexion)
    return usuario, alimentos

def exportar_a_pdf(usuario_id):
    usuario, alimentos = generar_reporte_usuario(usuario_id)
    if not usuario:
        print("❌ Usuario no encontrado")
        return
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Reporte Nutricional - NutriGest', 0, 1, 'C')
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"Usuario: {usuario['nombre']}", 0, 1)
    pdf.cell(0, 10, f"Edad: {usuario['edad']} años", 0, 1)
    pdf.cell(0, 10, f"Meta calórica diaria: {usuario['calorias_diarias']} kcal", 0, 1)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Últimos 15 alimentos consumidos:', 0, 1)
    pdf.set_font('Arial', '', 10)
    pdf.cell(100, 10, 'Alimento', 1)
    pdf.cell(30, 10, 'Cantidad', 1)
    pdf.cell(30, 10, 'Calorías', 1)
    pdf.cell(30, 10, 'Fecha', 1, 1)
    total_calorias = 0
    for alimento in alimentos:
        unidad = "g" if alimento['es_por_peso'] else "unid."
        pdf.cell(100, 10, alimento['nombre'], 1)
        pdf.cell(30, 10, f"{alimento['cantidad']}{unidad}", 1)
        pdf.cell(30, 10, f"{alimento['calorias_totales']:.2f} kcal", 1)
        pdf.cell(30, 10, str(alimento['fecha']), 1, 1)
        total_calorias += alimento['calorias_totales']
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(160, 10, 'Total calorías:', 1)
    pdf.cell(30, 10, f"{total_calorias:.2f} kcal", 1, 1)
    nombre_archivo = f"Reporte_Nutricional_{usuario_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    pdf.output(nombre_archivo)
    print(f"\n✅ Reporte PDF generado: {nombre_archivo}")

def exportar_a_excel(usuario_id):
    usuario, alimentos = generar_reporte_usuario(usuario_id)
    if not usuario:
        print("❌ Usuario no encontrado")
        return
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Reporte Nutricional"
    hoja.merge_cells('A1:D1')
    hoja['A1'] = "Reporte Nutricional - NutriGest"
    hoja['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    hoja['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    hoja['A2'] = "Usuario:"
    hoja['B2'] = usuario['nombre']
    hoja['A3'] = "Fecha:"
    hoja['B3'] = datetime.now().strftime('%d/%m/%Y')
    hoja['A5'] = "Alimento"
    hoja['B5'] = "Cantidad"
    hoja['C5'] = "Calorías (kcal)"
    hoja['D5'] = "Fecha"
    for celda in hoja[5]:
        celda.font = openpyxl.styles.Font(bold=True)
    fila = 6
    for alimento in alimentos:
        unidad = "g" if alimento['es_por_peso'] else "unid."
        hoja[f'A{fila}'] = alimento['nombre']
        hoja[f'B{fila}'] = f"{alimento['cantidad']} {unidad}"
        hoja[f'C{fila}'] = alimento['calorias_totales']
        hoja[f'D{fila}'] = str(alimento['fecha'])
        fila += 1
    for col in ['A','B','C','D']:
        hoja.column_dimensions[col].width = 20
    nombre_archivo = f"Reporte_Nutricional_{usuario_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    libro.save(nombre_archivo)
    print(f"\n✅ Reporte Excel generado: {nombre_archivo}")
